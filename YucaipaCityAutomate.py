
from PyPDF2 import PdfFileWriter, PdfFileReader, PdfFileMerger
import io, openpyxl, random
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

#takes inputs from excel input file, organizes it into dictionary d
def openpyxl():
    path = r'C:\Users\Lenovo\PycharmProjects\YucaipaCityPermit\input.xlsx'
    file = load_workbook(path)
    sheet = file.active

    d = {}
    for row in range(sheet.max_row):
        num = sheet.cell(row=row+1, column=2).value
        field = sheet.cell(row=row+1, column=3).value
        if type(num)==int  and field !=None:
            raw_input = sheet.cell(row=row + 1, column=4).value
            d[field] = raw_input

    #append a bunch of (hard coded) stuff to the back of the dictionary
    d = append_dic(d)
    print(d)
    return d

#append (hard coding) stuff to the excel-made-dictionary (above)
#ex)from "9_date", we can deduce that the month is "April" and day is "15"; add this stuff to the back of the dictionary
def append_dic(d):
    # todays_month = d['9_date'][0:4]
    d['Address'] = d['Address2']+', Yucaipa CA 92399'
    d['Address3'] = d['Address2'][:d['Address2'].index('Space')]+'Yucaipa CA 92399'
    d['Valuation'] = random.randint(8,18)*500
    d['work_description_line1'] = 'Mobile Home Set Down Space #'+str(d['SpaceNum'])
    d['work_description_line2'] = '('+str(d['Width'])+' x '+str(d['Length'])+' ft long)'
    d['Building_SF']= round(d['Width'] * d['Length'],0)
    d['ParkName'] = d['ParkName']+' Mobile Home Park'
    avgleftright = ((d['lot_right_len']+d['lot_left_len'])/2)
    avgupdown = ((d['lot_up_len']+d['lot_down_len'])/2)
    d['lotarea'] = avgleftright*avgupdown
    d['occ_area'] = round(d['Building_SF']/d['lotarea'],2)
    d['LLC'] = llc(d['ParkName'])
    d['freewayexit'] = authroute(d['ParkName'],0)+' Exit'
    d['localexit'] = authroute(d['ParkName'],1)
    return d

#given a park name, return a hauling route
def authroute(parkname, index):
    d = {'Hitching Post':['Yucaipa Blvd','Left into Hitching Post after 4th Street'],
    'Crestview':['Live Oak','Oak Glen Rd -> Ave E --> 4th St'],
    'Westwind':['Live Oak','Oak Glen Rd -> Ave E --> 4th St'],
    'Holiday':['County Line','Left into Holiday Rancho after Calimesa Blvd'],
    'Wishing Well':['Live Oak','Oak Glen Rd -> Ave E --> 5th St'],
    'Patrician':['County Line','Left into Patrician after 5th St'],
    'Mt Vista':['Live Oak','Oak Glen Rd -> Ave E --> 2nd St']
         }
    for i in d:
        if i in parkname:
            return d[i][index]
    return 'N/A'

#given a park name, return the corresponding LLC
def llc(parkname):
    d = {'Hitching Post':'Hitching Post Mobile Home Park, LLC',
    'Crestview':'Yucaipa Crestview, LLC',
    'Westwind':'Yucaipa Westwind Estates, LLC',
    'Holiday':'Holiday Rancho Park, LLC',
    'Wishing Well':'Wishing Well Mobile Home Park, LLC',
    'Patrician':'Patrician Mobile Home Park',
    'Mt Vista':'Mount Vista, LLC'
    }
    for i in d:
        if i in parkname:
            return d[i]
    return 'N/A'

#ink autocad PDF drawing with those sweet, sweet labels on bottom left hand corner
def ink_drawing():
    d = openpyxl()
    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=letter)
    can.drawString(10, 100, d['LLC'])
    can.drawString(10, 85, d['Address'])
    can.drawString(10, 70, "Owner - Jian Chen")
    can.drawString(10, 55, "Phone - 909 210 1491")
    s1 = "Footprint-- "+str(d['Width'])+' x '+str(d['Length'])+'-- '+str(d['Building_SF'])+'SF'
    s2 = "Lot-- "+str(d['lotarea'])+ " SF"
    s3 = str(d['Building_SF'])+" / "+ str(d['lotarea'])+' = '+str(d['occ_area'])
    can.drawString(10, 40, s1)
    can.drawString(10, 25, s2)
    can.drawString(10, 10, s3)
    can.save()

    #move to the beginning of the StringIO buffer
    packet.seek(0)

    # create a new PDF with Reportlab
    new_pdf = PdfFileReader(packet)
    # read your existing PDF
    input_path = r'C:\Users\Lenovo\PycharmProjects\YucaipaCityPermit\input\drawing.pdf'
    output_path = r'C:\Users\Lenovo\PycharmProjects\YucaipaCityPermit\output\drawingoutput.pdf'

    existing_pdf = PdfFileReader(open(input_path, "rb"))
    output = PdfFileWriter()
    # add the "watermark" (which is the new pdf) on the existing page
    page = existing_pdf.getPage(0)
    page.mergePage(new_pdf.getPage(0))
    output.addPage(page)
    # finally, write "output" to a real file
    outputStream = open(output_path, "wb")
    output.write(outputStream)
    outputStream.close()

def alterpdf(emptypath,filledpath):
    d = openpyxl()
    reader = PdfFileReader(emptypath)
    writer = PdfFileWriter()
    page = reader.pages[0]
    fields = reader.getFields()
    writer.addPage(page)
    # Now you add your data to the forms!
    for x in d:
        writer.updatePageFormFieldValues(
            writer.getPage(0), {x: d[x]}
        )
    # write "output" to PyPDF2-output.pdf
    with open(filledpath, "wb") as output_stream:
        writer.write(output_stream)
#fill up them PDFs baby
def fill():
    L = ['page1', 'page2', 'page3', 'page4', \
         'page5','page6','page7','page8','page9']
    for i in L:
        emptypath = 'C:\\Users\\Lenovo\\PycharmProjects\\YucaipaCityPermit\\input\\'+ i +'.pdf'
        filledpath = 'C:\\Users\\Lenovo\\PycharmProjects\\YucaipaCityPermit\\output\\' + i +'.pdf'
        alterpdf(emptypath,filledpath)

#combine every file in the filled path
def combine():
    merger = PdfFileMerger()
    L = ['page1', 'page2', 'page3', 'page4', \
         'page5', 'page6', 'page7','page8','page9','drawingoutput','drawingoutput','drawingoutput']
    for i in L:
        file = 'C:\\Users\\Lenovo\\PycharmProjects\\YucaipaCityPermit\\output\\' + i +'.pdf'
        merger.append(PdfFileReader(open(file,'rb')))
    merger.write(r'C:\Users\Lenovo\PycharmProjects\YucaipaCityPermit\printme\combined.pdf')
    return None

fill()
ink_drawing()
combine()